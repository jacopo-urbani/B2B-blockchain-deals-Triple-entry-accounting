import openpyxl
# update buyer's financial statements
def update_excel_financial_statement(name, total_amount, budget_item):
    excel_data = []
    box_numbers = []

    # Define variable to load the workbook
    workbook = openpyxl.load_workbook(name)

    # Define variable to read the active sheet:
    worksheet = workbook['CE']

    # Iterate the loop to read the cell values
    for i in range(0, worksheet.max_row):
        for col in worksheet.iter_cols(1, worksheet.max_column):
            # elimitating useless data
            if col[i].value != None:
                # appending box numbers to list 'box numbers'
                box_numbers.append(col[i])
                # appending box data to list 'excel_data'
                excel_data.append(col[i].value)
    # verifying if debt is a profit & loss account voice
    if budget_item in excel_data:
        statement = 'CE'
        # finding voice in the financial statement to get the old data
        data_item = excel_data.index(budget_item) + 1
        old_data = excel_data[data_item]
        data_index = str(box_numbers[data_item])
        index = data_index.partition('.')[2][:-1]
        item_statement = 'Budget item successfully recorded to the profit & loss account'
    # knowing that voice is found in the balance sheet
    else:
        statement = 'SP'
        # opening and taking all data as above
        worksheet = workbook[statement]
        for i in range(0, worksheet.max_row):
            for col in worksheet.iter_cols(1, worksheet.max_column):
                if col[i].value != None:
                    box_numbers.append(col[i])
                    excel_data.append(col[i].value)
        if budget_item in excel_data:
            # finding voice in the financial statement to get the old data
            data_item = excel_data.index(budget_item) + 1
            old_data = excel_data[data_item]
            data_index = str(box_numbers[data_item])
            index = data_index.partition('.')[2][:-1]
            item_statement = 'Budget item successfully recorded to the balance sheet'
    # overwriting the data by summing the old data and the new one
    xfile = openpyxl.load_workbook(name)
    sheet = xfile[statement]
    sheet[index] = (old_data) + (total_amount)
    xfile.save(name)
    print(item_statement)
